"""
Utility helpers for exporting the computed route into an Excel workbook.

The format mimics the sample workbook under ``test/`` where each customer
occupies one block consisting of two logical columns:

- Column group B–G: textual summary of ordered items per customer.
- Column H: customer name followed by the address split across multiple lines.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from TruckRouteApp.util import resolve_asset_path

DEFAULT_TEMPLATE = resolve_asset_path("template.xlsx")
DEFAULT_HEADER_IMAGE = resolve_asset_path("header_logo.png")

_COLUMN_ITEM_START = 2  # Column B
_COLUMN_SECONDARY = 8  # Column H
_TOP_PADDING_ROWS = 4
_COLUMN_WIDTHS = {
    2: 6,   # pallets value
    3: 6,   # "Pal."
    4: 4,   # "x"
    5: 28,  # item name
    6: 6,   # ktn qty
    7: 6,   # "ktn"
    8: 40,  # customer / address block (increased from 32)
}
_DATA_START_ROW = _TOP_PADDING_ROWS + 1
_CLEAR_ROW_COUNT = 400  # generous default to wipe prior runs without touching images

_THIN_SIDE = Side(style="thin", color="000000")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def _can_edit_cell(cell, merged_ranges) -> bool:
    for merged in merged_ranges:
        if cell.coordinate in merged:
            return cell.row == merged.min_row and cell.column == merged.min_col
    return True


@dataclass(slots=True)
class RouteExcelItem:
    """Single product line destined for a customer in the export."""

    item_name: str
    pallets: float
    cartons_per_pallet: Optional[float] = None


@dataclass(slots=True)
class RouteExcelRow:
    """
    Representation of one stop within the exported workbook.

    Rows should be supplied in route order (first destination after the depot
    through the final destination). The exporter takes care of reversing the
    order so that the first destination appears at the bottom of the sheet.
    """

    customer_name: str
    address: Optional[str]
    items: Sequence[RouteExcelItem] = field(default_factory=tuple)
    customer_id: Optional[str] = None

    @property
    def total_pallets(self) -> float:
        return float(sum(item.pallets for item in self.items))


def _load_workbook_from_template(path: Path) -> Tuple[Workbook, Worksheet]:
    # Load the template workbook to preserve images and formatting
    workbook = load_workbook(path)
    worksheet = workbook.active
    if worksheet is None:
        worksheet = workbook.create_sheet("Route")
    else:
        worksheet.title = "Route"
    for idx, width in _COLUMN_WIDTHS.items():
        worksheet.column_dimensions[get_column_letter(idx)].width = width
    _clear_data_region(worksheet)
    return workbook, worksheet


def _add_header_image(worksheet: Worksheet, image_path: Path) -> None:
    """
    Add an image to the header area, starting at column E.
    The image will be positioned above the customer data table.
    """
    if not image_path.exists():
        return
    
    try:
        # Set row heights for the header area (rows 1-3) to be 3 times taller
        normal_row_height = 15  # Default row height in Excel
        header_row_height = normal_row_height * 3

        worksheet.row_dimensions[3].height = header_row_height # row 3
        
        # Create image object
        img = Image(str(image_path))

        # Position the image starting at column E, row 2
        img.anchor = 'E2'  # Start at column E, row 2

        # Scale the image to fit in the enlarged header area
        # Column E width from _COLUMN_WIDTHS (item name column width)
        column_e_width_chars = _COLUMN_WIDTHS.get(5, 40)  # Column 5 (E) has width 40 characters
        # Convert Excel character width to approximate pixels (Excel uses ~7 pixels per character)
        max_width = int(column_e_width_chars * 7)
        max_height = int(header_row_height * 3)  # Available height across 3 tall rows
        
        # Get original dimensions
        original_width = img.width
        original_height = img.height
        
        # Calculate scale factor to maintain aspect ratio
        width_scale = max_width / original_width
        height_scale = max_height / original_height
        scale_factor = min(width_scale, height_scale)  # Use the smaller scale to fit within bounds
        
        # Apply proportional scaling with proper rounding
        img.width = round(original_width * scale_factor)
        img.height = round(original_height * scale_factor)
        
        # Add the image to the worksheet
        worksheet.add_image(img)
        
    except Exception as e:
        print(f"Warning: Could not add header image: {e}")


def _clear_data_region(worksheet: Worksheet) -> None:
    max_row = max(_DATA_START_ROW + _CLEAR_ROW_COUNT, worksheet.max_row or _DATA_START_ROW)
    merged = list(worksheet.merged_cells.ranges)
    for row in worksheet.iter_rows(
        min_row=_DATA_START_ROW,
        max_row=max_row,
        min_col=_COLUMN_ITEM_START,
        max_col=_COLUMN_SECONDARY,
    ):
        for cell in row:
            if _can_edit_cell(cell, merged):
                cell.value = None
                cell.border = Border()


def _apply_grid_border(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    last_customer_row: Optional[int] = None,
) -> None:
    """
    Draw a cleaner table border:
    - Outer border around the customer data region (not including summary rows)
    - Vertical borders only at: end of items section, and end of address column
    - Horizontal border only under final item of each customer block (via separate helper)
    """
    # If last_customer_row is specified and less than max_row, stop outer border at that row.
    border_end_row = last_customer_row if last_customer_row is not None and last_customer_row < max_row else max_row

    # Outer rectangle around the customer data (not summary)
    for col in range(min_col, max_col + 1):
        top = worksheet.cell(row=min_row, column=col)
        bottom = worksheet.cell(row=border_end_row, column=col)
        top.border = Border(top=_THIN_SIDE, left=top.border.left, right=top.border.right, bottom=top.border.bottom)
        bottom.border = Border(bottom=_THIN_SIDE, left=bottom.border.left, right=bottom.border.right, top=bottom.border.top)

    for row in range(min_row, border_end_row + 1):
        left = worksheet.cell(row=row, column=min_col)
        right = worksheet.cell(row=row, column=max_col)
        left.border = Border(left=_THIN_SIDE, top=left.border.top, right=left.border.right, bottom=left.border.bottom)
        right.border = Border(right=_THIN_SIDE, top=right.border.top, left=right.border.left, bottom=right.border.bottom)

    # Vertical border after items column group (ktn col)
    if last_customer_row is not None:
        divider_col = _COLUMN_ITEM_START + 5
        last_row_for_divider = last_customer_row
        for row in range(min_row, last_row_for_divider + 1):
            cell = worksheet.cell(row=row, column=divider_col)
            cell.border = Border(
                right=_THIN_SIDE,
                left=cell.border.left,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )
    # Do NOT draw vertical border in summary region (after last_customer_row)

    # Horizontal separator under full customer block (final row) handled separately

    # If there are summary rows after the last customer row, draw left/right border for those rows
    if border_end_row < max_row:
        for row in range(border_end_row + 1, max_row + 1):
            left = worksheet.cell(row=row, column=min_col)
            right = worksheet.cell(row=row, column=max_col)
            left.border = Border(left=_THIN_SIDE, top=left.border.top, right=left.border.right, bottom=left.border.bottom)
            right.border = Border(right=_THIN_SIDE, top=right.border.top, left=right.border.left, bottom=right.border.bottom)
def _apply_summary_border(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    """
    Draw a thin border around the summary rows (date and total), only for columns B and C.
    - Left border at column B, right border at column C, for all rows.
    - Top border at top row, bottom border at bottom row.
    - No vertical borders elsewhere, no border at other columns.
    """
    min_col = _COLUMN_ITEM_START
    max_col = _COLUMN_ITEM_START + 1
    for row in range(start_row, end_row + 1):
        left = worksheet.cell(row=row, column=min_col)
        right = worksheet.cell(row=row, column=max_col)
        # Determine top/bottom sides depending on whether this is the first/last row.
        left_top = _THIN_SIDE if row == start_row else left.border.top
        left_bottom = _THIN_SIDE if row == end_row else left.border.bottom
        right_top = _THIN_SIDE if row == start_row else right.border.top
        right_bottom = _THIN_SIDE if row == end_row else right.border.bottom
        # Always left/right border
        left_border = Border(
            left=_THIN_SIDE,
            top=left_top,
            right=left.border.right,
            bottom=left_bottom,
        )
        right_border = Border(
            right=_THIN_SIDE,
            top=right_top,
            left=right.border.left,
            bottom=right_bottom,
        )
        left.border = left_border
        right.border = right_border

def _apply_customer_separator(worksheet: Worksheet, row_idx: int) -> None:
    """Draw a horizontal line (bottom border) at row_idx across all columns of the main table."""
    # Draw a horizontal line (bottom border) at row_idx across all columns of the main table.
    for col in range(_COLUMN_ITEM_START, _COLUMN_SECONDARY + 1):
        cell = worksheet.cell(row=row_idx, column=col)
        cell.border = Border(
            bottom=_THIN_SIDE,
            top=cell.border.top,
            left=cell.border.left,
            right=cell.border.right,
        )


def _normalized_number(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    return round(float(value), 3)


def _split_address_lines(address: Optional[str]) -> list[str]:
    if not address:
        return []
    raw = address.strip()
    if not raw:
        return []
    if "," not in raw:
        return [raw]
    first, remainder = raw.split(",", 1)
    lines = [first.strip()]
    remainder = remainder.strip()
    if remainder:
        lines.append(remainder)
    return lines


def _write_item_line(worksheet: Worksheet, row_idx: int, item: RouteExcelItem) -> None:
    pallets_value = _normalized_number(item.pallets)
    cartons_value = _normalized_number(item.cartons_per_pallet)
    worksheet.cell(row=row_idx, column=_COLUMN_ITEM_START, value=pallets_value)
    worksheet.cell(row=row_idx, column=_COLUMN_ITEM_START + 1, value="Pal.")
    worksheet.cell(row=row_idx, column=_COLUMN_ITEM_START + 2, value="x")
    worksheet.cell(row=row_idx, column=_COLUMN_ITEM_START + 3, value=item.item_name)
    worksheet.cell(row=row_idx, column=_COLUMN_ITEM_START + 4, value=cartons_value)
    worksheet.cell(
        row=row_idx,
        column=_COLUMN_ITEM_START + 5,
        value="ktn" if cartons_value is not None else None,
    )


def _write_customer_block(worksheet: Worksheet, start_row: int, row: RouteExcelRow) -> tuple[int, Optional[int]]:
    """
    Write a customer block starting at start_row.
    Returns (next_row, separator_row) where separator_row is the row index where a horizontal line should be applied,
    or None if no items or address lines were written for this customer.
    """
    # Create horizontal center alignment for column H
    center_alignment = Alignment(horizontal='center')
    
    # Write customer name with center alignment
    customer_cell = worksheet.cell(row=start_row, column=_COLUMN_SECONDARY, value=row.customer_name)
    customer_cell.alignment = center_alignment
    
    current_row = start_row + 1
    address_lines = _split_address_lines(row.address)
    line_count = max(len(row.items), len(address_lines))
    if line_count == 0:
        return start_row, None

    for offset in range(line_count):
        if offset < len(row.items):
            _write_item_line(worksheet, current_row, row.items[offset])
        if offset < len(address_lines):
            # Write address line with center alignment
            address_cell = worksheet.cell(row=current_row, column=_COLUMN_SECONDARY, value=address_lines[offset])
            address_cell.alignment = center_alignment
        current_row += 1

    # Only add a blank padding row if there were items or address lines
    worksheet.cell(row=current_row, column=_COLUMN_SECONDARY, value=None)
    padding_row = current_row
    separator_row = padding_row  # separator goes at the padding row (bottom of customer block)
    return separator_row + 1, separator_row


def export_route_to_excel(
    output_path: Path,
    rows: Sequence[RouteExcelRow],
    *,
    order_date: Optional[datetime | date | str] = None,
    total_pallets: Optional[float] = None,
    template_path: Optional[Path] = None,
    header_image_path: Optional[Path] = None,
) -> Path:
    """
    Export the provided rows to an Excel file matching the sample layout.
    """

    template = template_path or DEFAULT_TEMPLATE
    workbook, worksheet = _load_workbook_from_template(template)

    current_row = _DATA_START_ROW
    table_start_row = current_row
    table_end_row = current_row - 1
    separator_rows = []
    for row in reversed(rows):
        next_row, separator_row = _write_customer_block(worksheet, current_row, row)
        if separator_row is not None:
            separator_rows.append(separator_row)
            table_end_row = max(table_end_row, separator_row)  # last row containing data (padding row)
            current_row = next_row
        # If separator_row is None, do not add phantom block or separator, do not advance current_row

    # Draw horizontal separator after each customer block
    for sep_row in separator_rows:
        _apply_customer_separator(worksheet, sep_row)

    last_customer_row = max(separator_rows) if separator_rows else table_end_row

    current_row += 1  # extra blank row before summary

    summary_start_row = current_row
    if order_date:
        if isinstance(order_date, datetime):
            date_value: str | datetime | date = order_date.strftime("%d.%m.%Y")
        elif isinstance(order_date, date):
            date_value = order_date.strftime("%d.%m.%Y")
        else:
            date_value = order_date
        worksheet.cell(row=current_row, column=_COLUMN_ITEM_START, value=date_value)
        # Merge columns B and C for the date row
        worksheet.merge_cells(start_row=summary_start_row, start_column=_COLUMN_ITEM_START,
                              end_row=summary_start_row, end_column=_COLUMN_ITEM_START + 1)
        current_row += 1

    total = _normalized_number(total_pallets) if total_pallets is not None else None
    if total is None:
        total = _normalized_number(sum(row.total_pallets for row in rows))
    worksheet.cell(row=current_row, column=_COLUMN_ITEM_START, value=total)
    worksheet.cell(row=current_row, column=_COLUMN_ITEM_START + 1, value="Pal.")
    summary_end_row = current_row

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if table_end_row >= table_start_row and rows:
        _apply_grid_border(
            worksheet,
            table_start_row,
            table_end_row,
            _COLUMN_ITEM_START,
            _COLUMN_SECONDARY,
            last_customer_row=last_customer_row,
        )
    if summary_end_row >= summary_start_row:
        _apply_summary_border(
            worksheet,
            summary_start_row,
            summary_end_row,
        )

    # Add header image only if explicitly provided (template already has the default image)
    if header_image_path and header_image_path.exists():
        _add_header_image(worksheet, header_image_path)
    
    workbook.save(output_path)
    return output_path


__all__ = ["RouteExcelItem", "RouteExcelRow", "export_route_to_excel", "DEFAULT_TEMPLATE", "DEFAULT_HEADER_IMAGE"]
